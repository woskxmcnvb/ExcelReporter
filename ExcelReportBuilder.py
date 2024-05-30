import io

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font


import pandas as pd

class ExcelReportBuilder:

    def __init__(self, file_name): 
        self.images_ = []
        self.tables_ = []
        self.file_name_ = file_name 

    def AddImage(self, stream: io.BytesIO, page_name: str, position: str="A1"):
        self.images_.append({'page': page_name, 'image': Image(stream), 'position': position})
    
    def AddTable(self, 
                 table: pd.DataFrame, 
                 page_name: str, 
                 title: None, 
                 description=None, 
                 drop_index=False, 
                 conditional_formatting=False):
        self.tables_.append(
            {
                'title': title, 
                'description': description,
                'page': page_name,
                'table': table, 
                'index': not(drop_index), 
                'conditional_formatting': conditional_formatting
            }
        )

    @staticmethod
    def __GetSheetPtr(wb: Workbook, name: str):
        if name not in wb.sheetnames:
            return wb.create_sheet(name)
        else: 
            return wb[name]

    
    def SaveToFile(self):
        wb = Workbook()
        
        #Writing tables
        for tab in self.tables_:
            ws = ExcelReportBuilder.__GetSheetPtr(wb, tab['page'])
            start_row = 1 if ws.max_row == 1 else ws.max_row + 2

            if tab['title']: 
                ws.cell(start_row, 1, tab['title']).font = Font(bold=True)
                start_row += 1

            if tab['description']: 
                ws.cell(start_row, 1, tab['description'])
                start_row += 1

            for r in dataframe_to_rows(tab['table'], index=tab['index'], header=True):
                ws.append(r)

            for r in range(start_row, ws.max_row+1):
                for cell in ws[r]:
                    cell.style = 'Pandas'

            if tab['conditional_formatting']:
                formatting_range = '{}{}:{}'.format(
                    ('A' if not tab['index'] else 'B'), 
                    start_row + 1, 
                    ws.cell(ws.max_row, ws.max_column).coordinate
                )
                ws.conditional_formatting.add(formatting_range, 
                    ColorScaleRule(
                        start_type='min', start_value=10, start_color='F8696B',
                        mid_type='percentile', mid_value=50, mid_color='FFEB84',
                        end_type='max', end_value=90, end_color='63BE7B'
                    ))
            ws.append([])

        # Writing images
        for img in self.images_:
            ws = ExcelReportBuilder.__GetSheetPtr(wb, img['page'])
            ws.add_image(img['image'], img['position'])

        del wb['Sheet']
        wb.save(self.file_name_)

